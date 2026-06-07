#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""보건복지부 전국 산후조리원 현황(반기) → price-national.js
   최신: 2025년 하반기(2025.12 기준) XLSX. 출처: mohw.go.kr board bid=0020 (로그인 불필요).
   다운로드: https://www.mohw.go.kr/boardDownload.es?bid=0020&list_no=<게시물>&seq=2 (seq=2=xlsx)
   실행: python build_price.py [xlsx경로]  (기본 sanhujo_price_202512.xlsx)
   조인키: phone=전화 digits, name=정규화상호|시군구. 단위 만원(2주). """
import sys, re, json
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "sanhujo_price_202512.xlsx"
LABEL = "2025.12"  # 기준일 표기


def pdig(t):
    return re.sub(r"\D", "", str(t or ""))


def pnorm(s):
    s = str(s or "")
    return re.sub(r"\s+", "", re.sub(r"\(.*?\)", "", re.sub(r"산후조리원", "", s))).lower()


def cprice(s):
    s = re.sub(r"[^0-9,]", "", str(s or "").strip())
    return s


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # 헤더 행 찾기 (번호/시도/... 포함)
    hi = 0
    for i, r in enumerate(rows[:10]):
        joined = "".join(str(c) for c in r if c)
        if "시도" in joined and "일반실" in joined:
            hi = i
            break
    by_phone, by_name = {}, {}
    n = 0
    for r in rows[hi + 1:]:
        if len(r) < 9:
            continue
        sido, sigungu, oper, name, addr, tel, normal, special = (
            str(r[1] or "").strip(), str(r[2] or "").strip(), r[3],
            str(r[4] or "").strip(), r[5], r[6], r[7], r[8])
        if not name or not sigungu:
            continue
        rec = {"n": cprice(normal), "s": cprice(special)}
        if not rec["n"] and not rec["s"]:
            continue
        d = pdig(tel)
        if d:
            by_phone[d] = rec
        by_name[pnorm(name) + "|" + sigungu] = rec
        n += 1
    out = {"y": LABEL, "phone": by_phone, "name": by_name}
    with open("price-national.js", "w", encoding="utf-8") as f:
        f.write("/* 전국 산후조리원 가격 (출처: 보건복지부 전국 산후조리원 현황 %s, 단위 만원, 2주 기준)\n" % LABEL)
        f.write("   조인키: phone=전화 digits, name=정규화상호|시군구. 자동 생성: build_price.py */\n")
        f.write("const PRICE_LATEST = " + json.dumps(out, ensure_ascii=False, separators=(",", ":")) + ";\n")
    print("생성: price-national.js | 기준 %s | 가격행 %d | by_phone %d | by_name %d" % (LABEL, n, len(by_phone), len(by_name)))


if __name__ == "__main__":
    main()
